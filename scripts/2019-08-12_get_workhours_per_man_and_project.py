#  WIP about finding time and material tasks within all companies and all tasks
#  will be the basis for the migration script, needs more testing

import os

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "systori.settings.local")
import django

django.setup()

from openpyxl import Workbook

from systori.apps.company.models import Company
from systori.apps.project.models import Project


if __name__ == "__main__":
    c = Company.objects.get(schema="mehr-handwerk")
    c.activate()
    print(f"ATTENTION - SPECIAL SCRIPT ONLY FOR SOFTRONIC (mehr-handwerk).\n")

    PROJECTS = [446, 383, 429, 502, 459, 489, 362, 481, 515, 516, 474, 451, 510]
    wb = Workbook()
    rohdaten = wb.create_sheet("rohdaten")
    workersWithDailyPlans = dict()
    for project in Project.objects.filter(id__in=PROJECTS):
        for jobsite in project.jobsites.all():
            for dailyplan in jobsite.dailyplans.all():
                for worker in dailyplan.workers.all():
                    workersWithDailyPlans.setdefault(worker, [])
                    workersWithDailyPlans[worker].append(
                        {
                            "date": dailyplan.day,
                            "project": project.name,
                            "project_id": project.id,
                            "working_hours": 8,
                        }
                    )
                    rohdaten.append(
                        [
                            dailyplan.day,
                            project.id,
                            project.name,
                            f"{worker.first_name} {worker.last_name}",
                            8,
                        ]
                    )

    projects = {"2018": {}, "2019": {}}
    for worker in workersWithDailyPlans:
        for i in range(len(workersWithDailyPlans[worker]) - 1, -1, -1):
            plan = workersWithDailyPlans[worker][i]
            if plan["date"].year == 2018:
                projects["2018"].setdefault(plan["project_id"], {})
                projects["2018"][plan["project_id"]].setdefault(
                    "project", plan["project"]
                )
                projects["2018"][plan["project_id"]].setdefault("dailyplans", [])
                projects["2018"][plan["project_id"]]["dailyplans"].append(plan)
                projects["2018"][plan["project_id"]].setdefault("total", 0)
                projects["2018"][plan["project_id"]]["total"] += plan["working_hours"]
                del workersWithDailyPlans[worker][i]
            elif plan["date"].year == 2019:
                projects["2019"].setdefault(plan["project_id"], {})
                projects["2019"][plan["project_id"]].setdefault(
                    "project", plan["project"]
                )
                projects["2019"][plan["project_id"]].setdefault("dailyplans", [])
                projects["2019"][plan["project_id"]]["dailyplans"].append(plan)
                projects["2019"][plan["project_id"]].setdefault("total", 0)
                projects["2019"][plan["project_id"]]["total"] += plan["working_hours"]
                del workersWithDailyPlans[worker][i]
            elif plan["date"].year == 2017:
                del workersWithDailyPlans[worker][i]
            else:
                raise KeyError

    uebersicht = wb.create_sheet("uebersicht")
    uebersicht.append(["2018"])
    for project in projects["2018"]:
        uebersicht.append(
            [projects["2018"][project]["project"], projects["2018"][project]["total"]]
        )
    uebersicht.append(["2019"])
    for project in projects["2019"]:
        uebersicht.append(
            [projects["2019"][project]["project"], projects["2019"][project]["total"]]
        )
    wb.save("test.xlsx")
